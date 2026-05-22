import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime, date   # 同时导入 date 类
import os
import re

def col_letter_to_index(letter):
    """将 Excel 列字母（如 A, Z, AA）转换为 1-based 列索引"""
    letter = letter.strip().upper()
    if not re.match(r'^[A-Z]+$', letter):
        raise ValueError(f"无效的列字母: {letter}")
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx

class CustomTableProcessor(ttk.LabelFrame):
    """1.自增加表格处理（可自定义保留列数量，Sheet2以保留列为重复列）"""
    def __init__(self, parent, **kwargs):
        super().__init__(parent, text="1.自增加表格处理", padding=10, **kwargs)
        self.processed_file = None
        self.column_entries = []
        self.create_widgets()

    def create_widgets(self):
        file_frame = ttk.Frame(self)
        file_frame.pack(fill="x", pady=5)
        ttk.Label(file_frame, text="原始数据文件:", font=('', 10, 'bold')).pack(side="left", padx=5)
        self.input_file_var = tk.StringVar()
        self.input_entry = ttk.Entry(file_frame, textvariable=self.input_file_var, width=40)
        self.input_entry.pack(side="left", padx=5)
        ttk.Button(file_frame, text="浏览", command=self.select_input_file).pack(side="left", padx=2)

        param_frame = ttk.LabelFrame(self, text="参数设置", padding=5)
        param_frame.pack(fill="x", pady=5)

        row0 = ttk.Frame(param_frame)
        row0.pack(fill="x", pady=2)
        ttk.Label(row0, text="保留列数 (A列起):").pack(side="left")
        self.col_count_var = tk.IntVar(value=3)
        col_count_spin = ttk.Spinbox(row0, from_=2, to=5, width=5, textvariable=self.col_count_var, command=self.on_col_count_change)
        col_count_spin.pack(side="left", padx=5)
        ttk.Label(row0, text="  (Sheet2将重复这些列)").pack(side="left", padx=5)

        self.dynamic_frame = ttk.Frame(param_frame)
        self.dynamic_frame.pack(fill="x", pady=5)

        row_data = ttk.Frame(param_frame)
        row_data.pack(fill="x", pady=2)
        ttk.Label(row_data, text="数据起始列号 (英文):").pack(side="left")
        self.start_col_var = tk.StringVar(value="J")
        ttk.Entry(row_data, textvariable=self.start_col_var, width=5).pack(side="left", padx=2)
        ttk.Label(row_data, text="每组列数:").pack(side="left", padx=10)
        self.group_size_var = tk.StringVar(value="9")
        ttk.Entry(row_data, textvariable=self.group_size_var, width=5).pack(side="left", padx=2)
        ttk.Label(row_data, text="连续空单元格数停止:").pack(side="left", padx=10)
        self.empty_threshold_var = tk.StringVar(value="5")
        ttk.Entry(row_data, textvariable=self.empty_threshold_var, width=5).pack(side="left", padx=2)

        self.build_dynamic_inputs()
        self.process_btn = ttk.Button(param_frame, text="处理数据 (生成中间文件)", command=self.process_data)
        self.process_btn.pack(pady=8)
        self.status_label = ttk.Label(self, text="未处理", foreground="gray")
        self.status_label.pack(anchor="w", pady=2)
        stats_frame = ttk.LabelFrame(self, text="数据处理统计", padding=5)
        stats_frame.pack(fill="x", pady=5)
        self.stats_text = tk.Text(stats_frame, height=3, width=70, state='disabled', wrap='word')
        self.stats_text.pack(fill='both', expand=True)

    def build_dynamic_inputs(self):
        for widget in self.dynamic_frame.winfo_children():
            widget.destroy()
        self.column_entries.clear()
        count = self.col_count_var.get()
        letters = ['A', 'B', 'C', 'D', 'E'][:count]
        default_vals = ['G', 'H', 'I', 'J', 'K'][:count]
        max_per_row = 3
        for i, (letter, default) in enumerate(zip(letters, default_vals)):
            row = i // max_per_row
            col = i % max_per_row
            frame = ttk.Frame(self.dynamic_frame)
            frame.grid(row=row, column=col, padx=10, pady=5, sticky='w')
            ttk.Label(frame, text=f"{letter}列来源列号:").pack(side="left")
            entry = ttk.Entry(frame, width=6)
            entry.insert(0, default)
            entry.pack(side="left", padx=5)
            self.column_entries.append(entry)
        if hasattr(self, 'start_col_var') and count > 0:
            last_default = default_vals[-1]
            last_idx = col_letter_to_index(last_default)
            next_idx = last_idx + 1
            def num_to_letter(num):
                if num <= 26:
                    return chr(ord('A') + num - 1)
                return last_default
            self.start_col_var.set(num_to_letter(next_idx))

    def on_col_count_change(self):
        self.build_dynamic_inputs()

    def select_input_file(self):
        file_path = filedialog.askopenfilename(
            title="选择原始数据Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.input_file_var.set(file_path)
            self.status_label.config(text="待处理", foreground="orange")
            self.processed_file = None
            self.stats_text.config(state='normal')
            self.stats_text.delete(1.0, tk.END)
            self.stats_text.config(state='disabled')

    def update_stats(self, sheet1_rows, sheet2_rows):
        self.stats_text.config(state='normal')
        self.stats_text.delete(1.0, tk.END)
        self.stats_text.insert(tk.END, f"Sheet1 处理数据行数（不含标题）: {sheet1_rows}\n")
        self.stats_text.insert(tk.END, f"Sheet2 生成数据行数（不含标题）: {sheet2_rows}\n")
        self.stats_text.config(state='disabled')

    def process_data(self):
        input_path = self.input_file_var.get().strip()
        if not input_path:
            messagebox.showerror("错误", "请先选择原始数据文件")
            return
        if not os.path.exists(input_path):
            messagebox.showerror("错误", "文件不存在")
            return
        try:
            col_numbers = [col_letter_to_index(entry.get().strip()) for entry in self.column_entries]
        except ValueError as e:
            messagebox.showerror("错误", f"保留列号输入无效：{str(e)}")
            return
        try:
            start_col_letter = self.start_col_var.get().strip()
            start_col = col_letter_to_index(start_col_letter)
            group_size = int(self.group_size_var.get().strip())
            empty_threshold = int(self.empty_threshold_var.get().strip())
        except ValueError as e:
            messagebox.showerror("错误", f"参数输入无效：{str(e)}")
            return
        if group_size <= 0 or empty_threshold <= 0:
            messagebox.showerror("错误", "组列数和空单元格阈值必须大于0")
            return
        output_path = os.path.join(os.path.dirname(input_path),
                                   f"{os.path.splitext(os.path.basename(input_path))[0]}_processed.xlsx")
        try:
            self.status_label.config(text="正在处理...", foreground="blue")
            self.update_idletasks()
            wb = load_workbook(input_path)
            if 'Sheet1' not in wb.sheetnames:
                raise ValueError("工作簿中没有名为 'Sheet1' 的工作表")
            ws1 = wb['Sheet1']
            for row in ws1.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "跳过" in cell.value:
                        cell.value = None
            if 'Sheet2' in wb.sheetnames:
                del wb['Sheet2']
            ws2 = wb.create_sheet('Sheet2')
            max_row = ws1.max_row
            max_col = ws1.max_column
            keep_headers = []
            for col_num in col_numbers:
                val = ws1.cell(row=1, column=col_num).value
                keep_headers.append(str(val).strip() if val is not None else f"列{col_num}")
            data_headers = []
            for i in range(group_size):
                col_idx = start_col + i
                if col_idx <= max_col:
                    header_val = ws1.cell(row=1, column=col_idx).value
                    data_headers.append(str(header_val).strip() if header_val is not None else "")
                else:
                    data_headers.append("")
            full_headers = keep_headers + data_headers
            for i, h in enumerate(full_headers, start=1):
                ws2.cell(row=1, column=i, value=h)
            output_row = 2
            sheet1_processed_rows = 0
            for r in range(2, max_row + 1):
                sheet1_processed_rows += 1
                keep_values = []
                for col_num in col_numbers:
                    keep_values.append(ws1.cell(row=r, column=col_num).value)
                data_row = []
                consecutive_none = 0
                col = start_col
                while True:
                    cell_val = ws1.cell(row=r, column=col).value if col <= max_col else None
                    if cell_val is None:
                        consecutive_none += 1
                        if consecutive_none == empty_threshold:
                            break
                        data_row.append(None)
                    else:
                        consecutive_none = 0
                        data_row.append(cell_val)
                    col += 1
                for i in range(0, len(data_row), group_size):
                    group = data_row[i:i+group_size]
                    if len(group) < group_size:
                        group += [None] * (group_size - len(group))
                    if all(v is None for v in group):
                        continue
                    for j, val in enumerate(keep_values, start=1):
                        ws2.cell(row=output_row, column=j, value=val)
                    for j, val in enumerate(group):
                        ws2.cell(row=output_row, column=len(keep_values) + 1 + j, value=val)
                    output_row += 1
            wb.save(output_path)
            wb.close()
            self.processed_file = output_path
            sheet2_generated_rows = output_row - 2
            self.update_stats(sheet1_processed_rows, sheet2_generated_rows)
            self.status_label.config(text=f"处理完成！中间文件: {os.path.basename(output_path)}", foreground="green")
            messagebox.showinfo("成功", f"数据处理完成！\n中间文件保存至：{output_path}\nSheet1 处理行数：{sheet1_processed_rows}\nSheet2 生成行数：{sheet2_generated_rows}")
        except Exception as e:
            self.status_label.config(text=f"处理失败: {str(e)}", foreground="red")
            messagebox.showerror("错误", f"数据处理失败：{str(e)}")


class OutputProcessor(ttk.LabelFrame):
    def __init__(self, parent, get_processed_file_func, **kwargs):
        super().__init__(parent, text="2. 输出文件格式与清理", padding=10, **kwargs)
        self.get_processed_file = get_processed_file_func
        self.date_column_entries = []
        self.float_column_entries = []
        self.create_widgets()

    def create_widgets(self):
        # 删除零值
        clean_frame = ttk.LabelFrame(self, text="删除零值", padding=5)
        clean_frame.pack(fill="x", pady=5)
        ttk.Label(clean_frame, text="一键清除 Sheet2 中所有值为 0 的单元格内容（仅清除，不删行）").pack(anchor="w")
        btn_clean = ttk.Button(clean_frame, text="删除所有零值单元格", command=self.delete_zero_cells)
        btn_clean.pack(pady=5)

        # 浮点格式
        float_frame = ttk.LabelFrame(self, text="设置浮点数字格式 (千位分隔符，小数最多4位)", padding=5)
        float_frame.pack(fill="x", pady=5)
        row2 = ttk.Frame(float_frame)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="列数量 (1-10):").pack(side="left")
        self.float_col_count = tk.IntVar(value=1)
        spin_float = ttk.Spinbox(row2, from_=1, to=10, width=3, textvariable=self.float_col_count, command=self.rebuild_float_inputs)
        spin_float.pack(side="left", padx=5)
        self.float_dynamic_frame = ttk.Frame(float_frame)
        self.float_dynamic_frame.pack(fill="x", pady=5)
        self.rebuild_float_inputs()
        btn_float = ttk.Button(float_frame, text="应用浮点格式", command=self.apply_float_format)
        btn_float.pack(pady=5)

        # 日期格式
        date_frame = ttk.LabelFrame(self, text="设置日期格式 (yyyy-mm-dd，自动转换文本为日期)", padding=5)
        date_frame.pack(fill="x", pady=5)
        row1 = ttk.Frame(date_frame)
        row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="列数量 (1-10):").pack(side="left")
        self.date_col_count = tk.IntVar(value=1)
        spin_date = ttk.Spinbox(row1, from_=1, to=10, width=3, textvariable=self.date_col_count, command=self.rebuild_date_inputs)
        spin_date.pack(side="left", padx=5)
        self.date_dynamic_frame = ttk.Frame(date_frame)
        self.date_dynamic_frame.pack(fill="x", pady=5)
        self.rebuild_date_inputs()
        btn_date = ttk.Button(date_frame, text="应用日期格式", command=self.apply_date_format)
        btn_date.pack(pady=5)

        self.status_label = ttk.Label(self, text="", foreground="gray")
        self.status_label.pack(anchor="w", pady=2)

    def rebuild_date_inputs(self):
        for widget in self.date_dynamic_frame.winfo_children():
            widget.destroy()
        self.date_column_entries.clear()
        count = self.date_col_count.get()
        max_per_row = 3
        for i in range(1, count+1):
            row = (i-1) // max_per_row
            col = (i-1) % max_per_row
            frame = ttk.Frame(self.date_dynamic_frame)
            frame.grid(row=row, column=col, padx=10, pady=5, sticky='w')
            ttk.Label(frame, text=f"第{i}列 (英文列号):").pack(side="left")
            entry = ttk.Entry(frame, width=6)
            entry.pack(side="left", padx=5)
            self.date_column_entries.append(entry)

    def rebuild_float_inputs(self):
        for widget in self.float_dynamic_frame.winfo_children():
            widget.destroy()
        self.float_column_entries.clear()
        count = self.float_col_count.get()
        max_per_row = 3
        for i in range(1, count+1):
            row = (i-1) // max_per_row
            col = (i-1) % max_per_row
            frame = ttk.Frame(self.float_dynamic_frame)
            frame.grid(row=row, column=col, padx=10, pady=5, sticky='w')
            ttk.Label(frame, text=f"第{i}列 (英文列号):").pack(side="left")
            entry = ttk.Entry(frame, width=6)
            entry.pack(side="left", padx=5)
            self.float_column_entries.append(entry)

    def _get_workbook_sheet2(self):
        file_path = self.get_processed_file()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("错误", "未找到中间文件，请先使用【1.自增加表格处理】生成文件")
            return None, None
        try:
            wb = load_workbook(file_path)
            if 'Sheet2' not in wb.sheetnames:
                messagebox.showerror("错误", "输出文件中没有 Sheet2 工作表")
                return None, None
            ws2 = wb['Sheet2']
            return wb, ws2
        except Exception as e:
            messagebox.showerror("错误", f"无法打开文件：{str(e)}")
            return None, None

    def apply_date_format(self):
        wb, ws2 = self._get_workbook_sheet2()
        if wb is None:
            return
        try:
            col_letters = [entry.get().strip() for entry in self.date_column_entries if entry.get().strip()]
            if not col_letters:
                messagebox.showwarning("警告", "未输入任何列号")
                return
            col_numbers = []
            for letter in col_letters:
                try:
                    col_numbers.append(col_letter_to_index(letter))
                except ValueError as e:
                    messagebox.showerror("错误", f"日期列号无效：{letter} -> {str(e)}")
                    return

            # 支持的日期格式（与Excel分列-日期YMD兼容）
            date_formats = [
                "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
                "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y"
            ]

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for col_idx in col_numbers:
                    if col_idx <= ws2.max_column:
                        cell = row[col_idx-1]
                        if cell.value is not None:
                            # 如果已经是日期或日期时间对象
                            if isinstance(cell.value, (datetime, date)):
                                cell.number_format = 'yyyy-mm-dd'
                            elif isinstance(cell.value, str):
                                # 尝试解析文本日期
                                date_obj = None
                                for fmt in date_formats:
                                    try:
                                        date_obj = datetime.strptime(cell.value.strip(), fmt).date()
                                        break
                                    except ValueError:
                                        continue
                                if date_obj:
                                    cell.value = date_obj
                                    cell.number_format = 'yyyy-mm-dd'
            wb.save(self.get_processed_file())
            wb.close()
            self.status_label.config(text="日期格式已应用（文本已转日期）", foreground="green")
            messagebox.showinfo("成功", f"已将第 {', '.join(col_letters)} 列设置为日期格式 (yyyy-mm-dd)，文本自动转换")
        except Exception as e:
            messagebox.showerror("错误", f"设置日期格式失败：{str(e)}")
            self.status_label.config(text="设置失败", foreground="red")

    def apply_float_format(self):
        wb, ws2 = self._get_workbook_sheet2()
        if wb is None:
            return
        try:
            col_letters = [entry.get().strip() for entry in self.float_column_entries if entry.get().strip()]
            if not col_letters:
                messagebox.showwarning("警告", "未输入任何列号")
                return
            col_numbers = []
            for letter in col_letters:
                try:
                    col_numbers.append(col_letter_to_index(letter))
                except ValueError as e:
                    messagebox.showerror("错误", f"浮点列号无效：{letter} -> {str(e)}")
                    return

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for col_idx in col_numbers:
                    if col_idx <= ws2.max_column:
                        cell = row[col_idx-1]
                        if cell.value is not None:
                            try:
                                if isinstance(cell.value, str):
                                    clean = cell.value.replace(',', '').strip()
                                    num = float(clean)
                                elif isinstance(cell.value, (int, float)):
                                    num = float(cell.value)
                                else:
                                    continue
                                cell.value = num
                                if num.is_integer():
                                    cell.number_format = '#,##0'
                                else:
                                    # 四舍五入到4位小数，然后确定有效小数位数
                                    rounded = round(num, 4)
                                    str_num = f"{rounded:.10f}".rstrip('0')
                                    if '.' in str_num:
                                        decimal_places = len(str_num.split('.')[1])
                                    else:
                                        decimal_places = 0
                                    if decimal_places == 0:
                                        cell.number_format = '#,##0'
                                    else:
                                        cell.number_format = '#,##0.' + '#' * decimal_places
                            except (ValueError, TypeError):
                                pass
            wb.save(self.get_processed_file())
            wb.close()
            self.status_label.config(text="浮点格式已应用", foreground="green")
            messagebox.showinfo("成功", f"已将第 {', '.join(col_letters)} 列设置为浮点格式（千位分隔符，小数最多4位）")
        except Exception as e:
            messagebox.showerror("错误", f"设置浮点格式失败：{str(e)}")
            self.status_label.config(text="设置失败", foreground="red")

    def delete_zero_cells(self):
        wb, ws2 = self._get_workbook_sheet2()
        if wb is None:
            return
        try:
            deleted_count = 0
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for cell in row:
                    if cell.value == 0 or (isinstance(cell.value, str) and cell.value.strip() == '0'):
                        cell.value = None
                        deleted_count += 1
            wb.save(self.get_processed_file())
            wb.close()
            self.status_label.config(text=f"已删除 {deleted_count} 个零值单元格", foreground="green")
            messagebox.showinfo("完成", f"共清除 {deleted_count} 个值为 0 的单元格内容")
        except Exception as e:
            messagebox.showerror("错误", f"删除零值失败：{str(e)}")
            self.status_label.config(text="删除失败", foreground="red")


class StandaloneApp:
    def __init__(self, root):
        self.root = root
        self.root.title("自增加表格处理器")
        self.root.geometry("750x700")
        style = ttk.Style()
        style.configure('TLabelframe.Label', font=('', 10, 'bold'))

        main_canvas = tk.Canvas(root, borderwidth=0, highlightthickness=0)
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        self.content_frame = ttk.Frame(main_canvas)
        main_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")
        self.content_frame.bind("<Configure>", lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")))
        def _on_mousewheel(event):
            main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        self.processor = CustomTableProcessor(self.content_frame)
        self.processor.pack(fill="x", padx=10, pady=10)
        self.output_processor = OutputProcessor(self.content_frame, get_processed_file_func=lambda: self.processor.processed_file)
        self.output_processor.pack(fill="x", padx=10, pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = StandaloneApp(root)
    root.mainloop()